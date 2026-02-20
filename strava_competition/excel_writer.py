"""Excel writer for segment and distance outputs (reads live in excel_reader)."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
import threading
from os import PathLike
from typing import Any, Iterable, Sequence
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from .models import Runner
from .errors import ExcelFormatError
from .segment_aggregation import (
    build_segment_outputs,
    ResultsMapping as SegResultsMapping,
    FASTEST_SEC_COL,
    FASTEST_FMT_COL,
    FASTEST_DATE_COL,
    FASTEST_DISTANCE_COL,
    BIRTHDAY_ATTR,
    TIME_BONUS_ATTR,
)

SEGMENTS_SHEET = "Segment Series"
RUNNERS_SHEET = "Runners"
DISTANCE_SHEET = "Distance Series"
STRAVA_ID_COLUMN = "Strava ID"
# Bandit B105 false positive: this is an Excel column header, not a secret.
REFRESH_TOKEN_COLUMN = "Refresh Token"  # nosec B105
SEGMENT_TEAM_COLUMN = "Segment Series Team"
DISTANCE_TEAM_COLUMN = "Distance Series Team"
BIRTHDAY_COLUMN = "Birthday (dd-mmm)"
MAX_SHEET_NAME_LEN = 31
EXCEL_DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"

__all__ = [
    "ExcelFormatError",
    "write_results",
    "update_runner_refresh_tokens",
    "update_single_runner_refresh_token",
]

ResultsMapping = SegResultsMapping
SUMMARY_HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(patternType="solid", fgColor="FFFF00FF")
DISTANCE_HEADER_FILL = PatternFill(patternType="solid", fgColor="FFFF6600")
BIRTHDAY_FILL = PatternFill(patternType="solid", fgColor="FFFFD89C")
TIME_BONUS_FILL = PatternFill(patternType="solid", fgColor="FFC9EFEA")
BOTH_BONUS_FILL = PatternFill(patternType="solid", fgColor="FFEFB5EF")
HEADER_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


LOGGER = logging.getLogger(__name__)


def _assert_file_exists(path: str) -> None:
    if not Path(path).is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")


_REQUIRED_RUNNER_COLS = {
    "Name",
    STRAVA_ID_COLUMN,
    REFRESH_TOKEN_COLUMN,
    SEGMENT_TEAM_COLUMN,
    DISTANCE_TEAM_COLUMN,
}


_WORKBOOK_LOCK = threading.RLock()


PathInput = str | Path | PathLike[str]
DistanceWindowRows = Sequence[dict[str, Any]]


def _coerce_path(pathlike: PathInput) -> str:
    return str(Path(pathlike))


## Segment aggregation lives in segment_aggregation.build_segment_outputs


def _unique_sheet_name(base: str, used: set[str]) -> str:
    base = base[:MAX_SHEET_NAME_LEN]
    name = base
    i = 1
    while name in used:
        suffix = f"_{i}"
        name = base[: MAX_SHEET_NAME_LEN - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _autosize(ws: Worksheet) -> None:
    from .config import (
        EXCEL_AUTOSIZE_COLUMNS,
        EXCEL_AUTOSIZE_MAX_WIDTH,
        EXCEL_AUTOSIZE_MIN_WIDTH,
        EXCEL_AUTOSIZE_PADDING,
        EXCEL_AUTOSIZE_MAX_ROWS,
    )

    if not EXCEL_AUTOSIZE_COLUMNS:
        return
    try:
        if ws.max_row > EXCEL_AUTOSIZE_MAX_ROWS:
            return
        for col_cells in ws.columns:
            max_len = 0
            col_letter = getattr(col_cells[0], "column_letter", None)
            for cell in col_cells:
                val = cell.value
                if val is None:
                    continue
                length = len(str(val))
                if length > max_len:
                    max_len = length
            width = min(
                EXCEL_AUTOSIZE_MAX_WIDTH,
                max(EXCEL_AUTOSIZE_MIN_WIDTH, max_len + EXCEL_AUTOSIZE_PADDING),
            )
            if col_letter:
                ws.column_dimensions[col_letter].width = width
    except Exception as exc:  # pragma: no cover - autosize is best-effort
        LOGGER.debug("Autosize failed for sheet %s: %s", getattr(ws, "title", "?"), exc)


## Row construction handled in segment_aggregation


def _format_runner_birthday_column(df: pd.DataFrame) -> None:
    if BIRTHDAY_COLUMN not in df.columns:
        return

    def _format_cell(value: object) -> object:
        if pd.isna(value) or str(value).strip() == "":
            return None
        parsed: pd.Timestamp | None = None
        if isinstance(value, tuple) and len(value) == 2:
            month, day = value
            try:
                parsed = pd.Timestamp(datetime(2000, int(month), int(day)))
            except (TypeError, ValueError):
                parsed = None
        if parsed is None:
            parsed = pd.to_datetime(value, errors="coerce")
        if (parsed is None or pd.isna(parsed)) and isinstance(value, str):
            text = value.strip()
            if text:
                try:
                    parsed = pd.Timestamp(datetime.strptime(f"{text}-2000", "%d-%b-%Y"))
                except ValueError:
                    parsed = None
        if parsed is None or pd.isna(parsed):
            return value
        month_label = parsed.strftime("%b")
        return f"{int(parsed.day):02d}-{month_label}"

    df[BIRTHDAY_COLUMN] = df[BIRTHDAY_COLUMN].map(_format_cell)


def _write_segment_sheets(
    writer: pd.ExcelWriter,
    results: ResultsMapping,
    used_sheet_names: set[str],
    include_summary: bool,
) -> None:
    outputs = build_segment_outputs(results, include_summary=include_summary)
    for base_name, df in outputs:
        sheet_name = _unique_sheet_name(base_name, used_sheet_names)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = _get_worksheet(writer, sheet_name)
        if ws is None:
            continue
        _style_header_row(ws, 1, len(df.columns))
        _apply_bonus_fills(ws, df)
        summary_df = getattr(df, "attrs", {}).get("segment_summary")
        if isinstance(summary_df, pd.DataFrame) and not summary_df.empty:
            _append_segment_summary(ws, summary_df)
        _autosize(ws)


## Summary building handled via segment_aggregation


def _write_distance_sheets(
    writer: pd.ExcelWriter,
    distance_windows_results: Iterable[tuple[str, DistanceWindowRows]],
    used_sheet_names: set[str],
    log_prefix: str = "Wrote distance window sheet",
) -> None:
    from .config import (
        DISTANCE_CREATE_EMPTY_WINDOW_SHEETS,
        DISTANCE_ENFORCE_COLUMN_ORDER,
        DISTANCE_COLUMN_ORDER,
    )

    for sheet_base, rows in distance_windows_results:
        if not rows and not DISTANCE_CREATE_EMPTY_WINDOW_SHEETS:
            continue
        dfw = pd.DataFrame(rows)
        if DISTANCE_ENFORCE_COLUMN_ORDER and not dfw.empty:
            ordered = [c for c in DISTANCE_COLUMN_ORDER if c in dfw.columns]
            remaining = [c for c in dfw.columns if c not in ordered]
            if ordered:
                dfw = dfw[ordered + remaining]
        sheet_name = _unique_sheet_name(sheet_base, used_sheet_names)
        dfw.to_excel(writer, sheet_name=sheet_name, index=False)
        LOGGER.info(
            "%s: %s rows=%s (empty=%s)",
            log_prefix,
            sheet_name,
            len(rows),
            not bool(rows),
        )
        ws = _get_worksheet(writer, sheet_name)
        if ws is not None:
            _style_header_row(ws, 1, len(dfw.columns), fill=DISTANCE_HEADER_FILL)
            _autosize(ws)


def write_results(
    filepath: PathInput,
    results: ResultsMapping,
    include_summary: bool = True,
    distance_windows_results: Sequence[tuple[str, DistanceWindowRows]] | None = None,
) -> None:
    filepath = _coerce_path(filepath)
    with pd.ExcelWriter(
        filepath, engine="openpyxl", datetime_format=EXCEL_DATETIME_FORMAT
    ) as writer:
        used_sheet_names: set[str] = set()
        if results:
            _write_segment_sheets(writer, results, used_sheet_names, include_summary)
        else:
            pd.DataFrame({"Message": ["No results to display."]}).to_excel(
                writer, sheet_name="Summary", index=False
            )
            _autosize_ws = _get_worksheet(writer, "Summary")
            if _autosize_ws is not None:
                _autosize(_autosize_ws)
        if distance_windows_results:
            _write_distance_sheets(writer, distance_windows_results, used_sheet_names)
            for sn in used_sheet_names:
                ws = _get_worksheet(writer, sn)
                if ws is not None:
                    _autosize(ws)


def _get_worksheet(writer: pd.ExcelWriter, sheet_name: str) -> Worksheet | None:
    try:
        return writer.book[sheet_name]
    except Exception:
        return writer.sheets.get(sheet_name)


def _append_segment_summary(ws: Worksheet, summary_df: pd.DataFrame) -> None:
    if summary_df.empty:
        return
    ws.append([])
    ws.append([])
    header = list(summary_df.columns)
    ws.append(header)
    header_row_idx = ws.max_row
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = SUMMARY_HEADER_FONT
    _style_header_row(ws, header_row_idx, len(header))
    for row in summary_df.itertuples(index=False, name=None):
        ws.append(list(row))


def _style_header_row(
    ws: Worksheet,
    row_idx: int,
    max_col: int | None = None,
    fill: PatternFill | None = None,
) -> None:
    if row_idx <= 0:
        return
    max_col = max_col or ws.max_column
    fill = fill or HEADER_FILL
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.border = HEADER_BORDER


def _apply_bonus_fills(ws: Worksheet, df: pd.DataFrame) -> None:
    """Apply fill colors to cells based on birthday and/or time bonus.

    - Birthday bonus only: BIRTHDAY_FILL (peach)
    - Time bonus only: TIME_BONUS_FILL (mint)
    - Both bonuses: BOTH_BONUS_FILL (lavender)
    """
    if ws is None or df is None:
        return
    birthday_rows = set(getattr(df, "attrs", {}).get(BIRTHDAY_ATTR) or [])
    time_bonus_rows = set(getattr(df, "attrs", {}).get(TIME_BONUS_ATTR) or [])
    if not birthday_rows and not time_bonus_rows:
        return
    columns = list(df.columns)

    def _col_index(name: str) -> int | None:
        try:
            return columns.index(name) + 1
        except ValueError:
            return None

    target_cols = [
        _col_index(FASTEST_SEC_COL),
        _col_index(FASTEST_FMT_COL),
        _col_index(FASTEST_DATE_COL),
        _col_index(FASTEST_DISTANCE_COL),
    ]
    target_cols = [idx for idx in target_cols if idx is not None]
    if not target_cols:
        return

    all_rows = birthday_rows | time_bonus_rows
    for row_idx in all_rows:
        try:
            excel_row = int(row_idx) + 2  # account for header row
        except (TypeError, ValueError):
            continue
        has_birthday = row_idx in birthday_rows
        has_time_bonus = row_idx in time_bonus_rows
        if has_birthday and has_time_bonus:
            fill = BOTH_BONUS_FILL
        elif has_birthday:
            fill = BIRTHDAY_FILL
        else:
            fill = TIME_BONUS_FILL
        for col_idx in target_cols:
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = fill


def _normalise_value(value: object) -> str:
    """Normalise Strava IDs for consistent comparisons."""

    if value is None:
        return ""
    normalised = str(value).strip()
    if normalised.endswith(".0"):
        normalised = normalised[:-2]
    return normalised


def _normalise_ids(series: pd.Series) -> pd.Series:
    return series.map(_normalise_value)


def update_runner_refresh_tokens(
    filepath: PathInput, runners: Sequence[Runner]
) -> None:
    filepath = _coerce_path(filepath)
    _assert_file_exists(filepath)
    with _WORKBOOK_LOCK:
        df = pd.read_excel(filepath, sheet_name=RUNNERS_SHEET)
        missing = _REQUIRED_RUNNER_COLS - set(df.columns)
        if missing:
            raise ExcelFormatError(
                f"Missing columns in '{RUNNERS_SHEET}' sheet: {', '.join(sorted(missing))}"
            )
        normalised_ids = _normalise_ids(df[STRAVA_ID_COLUMN])
        for runner in runners:
            runner_id = _normalise_value(runner.strava_id)
            mask = normalised_ids == runner_id
            if not mask.any():
                continue
            df.loc[mask, REFRESH_TOKEN_COLUMN] = runner.refresh_token
        _format_runner_birthday_column(df)
        with pd.ExcelWriter(
            filepath, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=RUNNERS_SHEET, index=False)


def update_single_runner_refresh_token(filepath: PathInput, runner: Runner) -> None:
    """Persist refresh token for a single runner (crash-safe incremental update).

    Reads only the Runners sheet, updates the row for the given runner, rewrites
    that sheet. Lightweight enough for occasional rotations.
    """
    filepath = _coerce_path(filepath)
    _assert_file_exists(filepath)
    with _WORKBOOK_LOCK:
        try:
            df = pd.read_excel(filepath, sheet_name=RUNNERS_SHEET)
        except Exception:
            return
        if STRAVA_ID_COLUMN not in df.columns or REFRESH_TOKEN_COLUMN not in df.columns:
            return
        id_series = _normalise_ids(df[STRAVA_ID_COLUMN])
        runner_id = _normalise_value(runner.strava_id)
        df.loc[id_series == runner_id, REFRESH_TOKEN_COLUMN] = runner.refresh_token
        _format_runner_birthday_column(df)
        try:
            with pd.ExcelWriter(
                filepath, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=RUNNERS_SHEET, index=False)
        except Exception as exc:
            LOGGER.warning(
                "Failed to persist refresh token for runner %s: %s",
                runner.name,
                exc,
            )
            return
