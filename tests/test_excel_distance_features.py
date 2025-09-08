import pandas as pd
from pathlib import Path
from strava_competition.excel_writer import write_results
from strava_competition.models import SegmentResult
from strava_competition import config


def build_segment_results():
    # minimal single segment, two teams
    r1 = SegmentResult(runner="A", team="Red", segment="Seg1", attempts=2, fastest_time=100.0, fastest_date=pd.Timestamp("2025-01-01"))
    r2 = SegmentResult(runner="B", team="Blue", segment="Seg1", attempts=1, fastest_time=95.0, fastest_date=pd.Timestamp("2025-01-02"))
    return {"Seg1": {"Red": [r1], "Blue": [r2]}}


def test_distance_sheet_column_order_and_empty(tmp_path, monkeypatch):
    # Force creation of empty distance sheet
    monkeypatch.setenv("DISTANCE_CREATE_EMPTY", "1")
    # Patch config flags dynamically (module variables already imported)
    from importlib import reload
    from strava_competition import config as cfg
    cfg.DISTANCE_CREATE_EMPTY_WINDOW_SHEETS = True
    cfg.DISTANCE_ENFORCE_COLUMN_ORDER = True

    distance_results = [
        ("Distance_2025-01-01_2025-01-07", [
            {"Runner": "A", "Team": "Red", "Runs": 3, "Total Distance (km)": 12.34, "Total Elev Gain (m)": 120.0},
            {"Runner": "B", "Team": "Blue", "Runs": 2, "Total Distance (km)": 10.0, "Total Elev Gain (m)": 90.0},
        ]),
        ("Distance_EmptyWindow", []),  # should still be created due to flag
    ]

    out_file = tmp_path / "out.xlsx"
    write_results(out_file, build_segment_results(), distance_windows_results=distance_results)

    import openpyxl
    wb = openpyxl.load_workbook(out_file)
    assert "Distance_2025-01-01_2025-01-07" in wb.sheetnames
    assert "Distance_EmptyWindow" in wb.sheetnames  # created despite empty

    # Verify column ordering (first distance sheet)
    ws = wb["Distance_2025-01-01_2025-01-07"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Expected ordered prefix
    assert header[:5] == ["Runner", "Team", "Runs", "Total Distance (km)", "Total Elev Gain (m)"]
