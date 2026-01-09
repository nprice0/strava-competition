from datetime import datetime
import os
import tempfile

import pandas as pd
import pytest
from openpyxl import load_workbook

from strava_competition import excel_writer, excel_reader
from strava_competition.excel_writer import (
    BIRTHDAY_FILL,
    TIME_BONUS_FILL,
    BOTH_BONUS_FILL,
)
from strava_competition.services.segment_service import SegmentService


def make_input_workbook(path):
    import pandas as pd

    segs = pd.DataFrame(
        [
            {
                "Segment ID": 101,
                "Segment Name": "Hill Climb",
                "Start Date": datetime(2024, 1, 1),
                "End Date": datetime(2024, 1, 31),
                "Default Time": None,
                "Minimum Distance (m)": 0,
                "Birthday Bonus (secs)": 5,
            }
        ]
    )
    runners = pd.DataFrame(
        [
            {
                "Name": "Alice",
                "Strava ID": 1,
                "Refresh Token": "rt1",
                "Segment Series Team": "Red",
                "Distance Series Team": "Red",
                "Birthday (dd-mmm)": "15-Jan",
            },
            {
                "Name": "Ben",
                "Strava ID": 2,
                "Refresh Token": "rt2",
                "Segment Series Team": "Blue",
                "Distance Series Team": None,
                "Birthday (dd-mmm)": "01-Feb",
            },
            {
                "Name": "Carl",
                "Strava ID": 3,
                "Refresh Token": "rt3",
                "Segment Series Team": "Red",
                "Distance Series Team": "Red",
                "Birthday (dd-mmm)": "20-Mar",
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        segs.to_excel(w, sheet_name="Segment Series", index=False)
        runners.to_excel(w, sheet_name="Runners", index=False)


@pytest.mark.filterwarnings("ignore::DeprecationWarning")
def test_excel_integration_roundtrip_and_ranks(monkeypatch):
    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, "input.xlsx")
        out_path = os.path.join(tmpdir, "out.xlsx")
        make_input_workbook(in_path)
        segments = excel_reader.read_segments(in_path)
        runners = excel_reader.read_runners(in_path)

        def fake_get_efforts(runner, segment_id, start_date, end_date):
            data = {
                "Alice": [
                    {"elapsed_time": 120, "start_date_local": "2024-01-10T09:00:00Z"},
                    {"elapsed_time": 115, "start_date_local": "2024-01-15T09:00:00Z"},
                ],
                "Ben": [
                    {"elapsed_time": 118, "start_date_local": "2024-01-12T09:00:00Z"},
                ],
                "Carl": [
                    {"elapsed_time": 115, "start_date_local": "2024-01-14T09:00:00Z"},
                    {"elapsed_time": 112, "start_date_local": "2024-01-20T09:00:00Z"},
                ],
            }
            return data.get(runner.name, [])

        import strava_competition.services.segment_service as segment_service_mod

        def fake_get_activities(runner, start_date, end_date):
            return []

        monkeypatch.setattr(
            segment_service_mod, "get_segment_efforts", fake_get_efforts
        )
        monkeypatch.setattr(segment_service_mod, "get_activities", fake_get_activities)
        monkeypatch.setattr(
            segment_service_mod, "FORCE_ACTIVITY_SCAN_FALLBACK", False, raising=False
        )

        service = SegmentService(max_workers=2)
        results = service.process(segments, runners)
        excel_writer.write_results(out_path, results)
        df = pd.read_excel(out_path, sheet_name="Hill Climb")
        wb = load_workbook(out_path, data_only=True)
        ws = wb["Hill Climb"]

    assert {
        "Team",
        "Runner",
        "Rank",
        "Team Rank",
        "Attempts",
        "Fastest Time (sec)",
        "Fastest Time (h:mm:ss)",
        "Fastest Date",
        "Fastest Distance (m)",
    }.issubset(set(df.columns))
    # Raw sheet includes blank spacer and summary rows - drop them for runner-level assertions.
    assert df["Team"].isna().sum() >= 1
    summary_markers = df.index[df["Team"] == "Team"]
    assert len(summary_markers) == 1
    summary_start = summary_markers[0]
    assert summary_start >= 2
    assert df.loc[summary_start - 2 : summary_start - 1, "Team"].isna().all()
    detail_df = (
        df.loc[: summary_start - 1].dropna(subset=["Runner"]).reset_index(drop=True)
    )
    assert len(detail_df) == 3
    df_sorted = detail_df.sort_values("Runner").reset_index(drop=True)
    runner_to_rank = dict(zip(df_sorted["Runner"], df_sorted["Rank"]))
    assert runner_to_rank["Alice"] == 1
    assert runner_to_rank["Carl"] == 2
    assert runner_to_rank["Ben"] == 3
    runner_to_time = dict(zip(df_sorted["Runner"], df_sorted["Fastest Time (h:mm:ss)"]))
    assert runner_to_time["Alice"] == "0:01:50"
    # Segment summary table appended beneath detail rows
    summary_header_row = None
    for idx in range(2, ws.max_row + 1):
        if ws.cell(row=idx, column=1).value == "Team" and ws.cell(
            row=idx - 1, column=1
        ).value in (None, ""):
            summary_header_row = idx
            break
    assert summary_header_row is not None
    assert ws.cell(row=summary_header_row - 1, column=1).value in (None, "")
    assert ws.cell(row=summary_header_row - 2, column=1).value in (None, "")
    assert ws.cell(row=summary_header_row, column=1).font.bold is True
    summary_rows = []
    row_idx = summary_header_row + 1
    while row_idx <= ws.max_row:
        team_cell = ws.cell(row=row_idx, column=1).value
        if not team_cell:
            break
        summary_rows.append(
            [ws.cell(row=row_idx, column=col).value for col in range(1, 10)]
        )
        row_idx += 1
    wb.close()
    assert len(summary_rows) == 2
    assert summary_rows[0][0] == "Blue"
    assert summary_rows[0][1] == "0:01:58"
    assert summary_rows[0][2] == "0:00:00"
    assert summary_rows[0][3] == 3
    assert summary_rows[1][0] == "Red"
    assert summary_rows[1][1] == "0:03:42"
    assert summary_rows[1][2] == "0:01:44"
    assert summary_rows[1][3] == 3
    assert summary_rows[1][4].startswith("Alice")
    assert float(summary_rows[0][-1]) == pytest.approx(1.0)
    assert float(summary_rows[1][-1]) == pytest.approx(2.0)

    headers = [cell.value for cell in ws[1]]
    fastest_sec_idx = headers.index("Fastest Time (sec)") + 1
    fastest_fmt_idx = headers.index("Fastest Time (h:mm:ss)") + 1
    fastest_date_idx = headers.index("Fastest Date") + 1
    alice_row_idx = None
    for row_idx in range(2, summary_header_row):
        if ws.cell(row=row_idx, column=2).value == "Alice":
            alice_row_idx = row_idx
            break
    assert alice_row_idx is not None
    for col_idx in (fastest_sec_idx, fastest_fmt_idx, fastest_date_idx):
        fill = ws.cell(row=alice_row_idx, column=col_idx).fill
        assert fill is not None
        # Verify birthday fill is applied using the constant (not hardcoded color)
        assert fill.fgColor.rgb == BIRTHDAY_FILL.fgColor.rgb


@pytest.mark.filterwarnings("ignore::DeprecationWarning")
def test_time_bonus_fill_applied(monkeypatch):
    """Verify TIME_BONUS_FILL is applied to runners with time bonus only."""
    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, "input.xlsx")
        out_path = os.path.join(tmpdir, "out.xlsx")

        # Create workbook with time bonus but no birthday bonus
        segs = pd.DataFrame(
            [
                {
                    "Segment ID": 101,
                    "Segment Name": "Hill Climb",
                    "Start Date": datetime(2024, 1, 1),
                    "End Date": datetime(2024, 1, 31),
                    "Default Time": None,
                    "Minimum Distance (m)": 0,
                    "Birthday Bonus (secs)": 0,  # No birthday bonus
                    "Time Bonus (secs)": 10,  # Time bonus enabled
                }
            ]
        )
        runners = pd.DataFrame(
            [
                {
                    "Name": "Bob",
                    "Strava ID": 2,
                    "Refresh Token": "rt2",
                    "Segment Series Team": "Blue",
                    "Distance Series Team": None,
                    "Birthday (dd-mmm)": "01-Jul",  # Birthday NOT in window
                },
            ]
        )
        with pd.ExcelWriter(in_path, engine="openpyxl") as w:
            segs.to_excel(w, sheet_name="Segment Series", index=False)
            runners.to_excel(w, sheet_name="Runners", index=False)

        segments = excel_reader.read_segment_groups(in_path)
        runners_list = excel_reader.read_runners(in_path)

        def fake_get_efforts(runner, segment_id, start_date, end_date):
            return [{"elapsed_time": 100, "start_date_local": "2024-01-15T09:00:00Z"}]

        import strava_competition.services.segment_service as segment_service_mod

        monkeypatch.setattr(
            segment_service_mod, "get_segment_efforts", fake_get_efforts
        )
        monkeypatch.setattr(segment_service_mod, "get_activities", lambda *a, **k: [])
        monkeypatch.setattr(segment_service_mod, "FORCE_ACTIVITY_SCAN_FALLBACK", False)
        monkeypatch.setattr(segment_service_mod, "SEGMENT_SPLIT_WINDOWS_ENABLED", True)

        service = SegmentService(max_workers=1)
        results = service.process_groups(segments, runners_list)
        excel_writer.write_results(out_path, results)

        wb = load_workbook(out_path, data_only=True)
        ws = wb["Hill Climb"]

        # Find Bob's row and verify time bonus fill is applied
        headers = [cell.value for cell in ws[1]]
        fastest_sec_idx = headers.index("Fastest Time (sec)") + 1
        bob_row_idx = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=2).value == "Bob":
                bob_row_idx = row_idx
                break

        assert bob_row_idx is not None
        fill = ws.cell(row=bob_row_idx, column=fastest_sec_idx).fill
        assert fill is not None
        assert fill.fgColor.rgb == TIME_BONUS_FILL.fgColor.rgb
        wb.close()


@pytest.mark.filterwarnings("ignore::DeprecationWarning")
def test_both_bonus_fill_applied(monkeypatch):
    """Verify BOTH_BONUS_FILL is applied when birthday and time bonus both apply."""
    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, "input.xlsx")
        out_path = os.path.join(tmpdir, "out.xlsx")

        # Create workbook with both bonuses
        segs = pd.DataFrame(
            [
                {
                    "Segment ID": 101,
                    "Segment Name": "Hill Climb",
                    "Start Date": datetime(2024, 1, 1),
                    "End Date": datetime(2024, 1, 31),
                    "Default Time": None,
                    "Minimum Distance (m)": 0,
                    "Birthday Bonus (secs)": 5,  # Birthday bonus enabled
                    "Time Bonus (secs)": 10,  # Time bonus enabled
                }
            ]
        )
        runners = pd.DataFrame(
            [
                {
                    "Name": "Carol",
                    "Strava ID": 3,
                    "Refresh Token": "rt3",
                    "Segment Series Team": "Green",
                    "Distance Series Team": None,
                    "Birthday (dd-mmm)": "15-Jan",  # Birthday IN window
                },
            ]
        )
        with pd.ExcelWriter(in_path, engine="openpyxl") as w:
            segs.to_excel(w, sheet_name="Segment Series", index=False)
            runners.to_excel(w, sheet_name="Runners", index=False)

        segments = excel_reader.read_segment_groups(in_path)
        runners_list = excel_reader.read_runners(in_path)

        def fake_get_efforts(runner, segment_id, start_date, end_date):
            # Effort on Carol's birthday
            return [{"elapsed_time": 100, "start_date_local": "2024-01-15T09:00:00Z"}]

        import strava_competition.services.segment_service as segment_service_mod

        monkeypatch.setattr(
            segment_service_mod, "get_segment_efforts", fake_get_efforts
        )
        monkeypatch.setattr(segment_service_mod, "get_activities", lambda *a, **k: [])
        monkeypatch.setattr(segment_service_mod, "FORCE_ACTIVITY_SCAN_FALLBACK", False)
        monkeypatch.setattr(segment_service_mod, "SEGMENT_SPLIT_WINDOWS_ENABLED", True)

        service = SegmentService(max_workers=1)
        results = service.process_groups(segments, runners_list)
        excel_writer.write_results(out_path, results)

        wb = load_workbook(out_path, data_only=True)
        ws = wb["Hill Climb"]

        # Find Carol's row and verify both-bonus fill is applied
        headers = [cell.value for cell in ws[1]]
        fastest_sec_idx = headers.index("Fastest Time (sec)") + 1
        carol_row_idx = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=2).value == "Carol":
                carol_row_idx = row_idx
                break

        assert carol_row_idx is not None
        fill = ws.cell(row=carol_row_idx, column=fastest_sec_idx).fill
        assert fill is not None
        assert fill.fgColor.rgb == BOTH_BONUS_FILL.fgColor.rgb
        wb.close()


@pytest.mark.filterwarnings("ignore::DeprecationWarning")
def test_time_bonus_applied_field_set(monkeypatch):
    """Verify time_bonus_applied field is set on SegmentResult."""
    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, "input.xlsx")

        segs = pd.DataFrame(
            [
                {
                    "Segment ID": 101,
                    "Segment Name": "Hill Climb",
                    "Start Date": datetime(2024, 1, 1),
                    "End Date": datetime(2024, 1, 31),
                    "Default Time": None,
                    "Minimum Distance (m)": 0,
                    "Birthday Bonus (secs)": 0,
                    "Time Bonus (secs)": 15,
                }
            ]
        )
        runners = pd.DataFrame(
            [
                {
                    "Name": "Dave",
                    "Strava ID": 4,
                    "Refresh Token": "rt4",
                    "Segment Series Team": "Red",
                    "Distance Series Team": None,
                    "Birthday (dd-mmm)": "01-Dec",
                },
            ]
        )
        with pd.ExcelWriter(in_path, engine="openpyxl") as w:
            segs.to_excel(w, sheet_name="Segment Series", index=False)
            runners.to_excel(w, sheet_name="Runners", index=False)

        segments = excel_reader.read_segment_groups(in_path)
        runners_list = excel_reader.read_runners(in_path)

        def fake_get_efforts(runner, segment_id, start_date, end_date):
            return [{"elapsed_time": 100, "start_date_local": "2024-01-10T09:00:00Z"}]

        import strava_competition.services.segment_service as segment_service_mod

        monkeypatch.setattr(
            segment_service_mod, "get_segment_efforts", fake_get_efforts
        )
        monkeypatch.setattr(segment_service_mod, "get_activities", lambda *a, **k: [])
        monkeypatch.setattr(segment_service_mod, "FORCE_ACTIVITY_SCAN_FALLBACK", False)
        monkeypatch.setattr(segment_service_mod, "SEGMENT_SPLIT_WINDOWS_ENABLED", True)

        service = SegmentService(max_workers=1)
        results = service.process_groups(segments, runners_list)

        dave_result = results["Hill Climb"]["Red"][0]
        # Verify the field is set directly on SegmentResult
        assert dave_result.time_bonus_applied is True
        # Verify adjusted time: 100 - 15 = 85
        assert dave_result.fastest_time == 85.0
